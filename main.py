import streamlit as st
import pandas as pd
from decouple import config
import requests
from bs4 import BeautifulSoup
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font

API_TOKEN  = config("TOKEN")
CANVAS_API_URL = 'https://canvas.uautonoma.cl/api/v1/'
headers = {
    'Authorization': f'Bearer {API_TOKEN }'
}

st.set_page_config(layout="wide",page_title="Super Course Checker", page_icon="🚀")
# Lista de subcuentas (IDs y nombres descriptivos)
subcuentas = {
    "Especialidad de Endodoncia": 746,
    "Especialidad de Rehabilitación Oral": 734,
    "Especialidad de Implantología Buco Maxilofacial": 732,
    "Especialidad Odontológica en Imagenología Oral y Maxilofacial": 459,
    "Especialidad en Medicina Familiar": 745,
    "Especialidad Médica en Medicina Interna": 743,
    "Especialidad en Imagenología Medica": 748,
    "Especialidad en Medicina de Urgencia": 747,
}

opciones_subcuentas = ["Seleccione una Especialidad"] + list(subcuentas.keys())
subcuenta_seleccionada = st.selectbox("Selecciona la Especialidad que deseas revisar:", opciones_subcuentas)


def obtener_cursos(subcuenta_id):
    cursos = []
    url = f"{CANVAS_API_URL}accounts/{subcuenta_id}/courses"
    params = {
        "per_page": 100,
        "include[]": ["sis_course_id"]
    }

    while url:
        response = requests.get(url, headers=headers, params=params)
        if response.status_code != 200:
            st.error(f"Error al obtener los cursos: {response.status_code} - {response.text}")
            return []
        try:
            data = response.json()
            cursos.extend([curso for curso in data if not curso.get("blueprint") and not '2022' in curso.get('sis_course_id')])
            # Obtener el siguiente enlace de la paginación
            url = response.links.get('next', {}).get('url')
        except ValueError:
            st.error("La respuesta de la API no es un JSON válido.")
            return []
        params = None  # Después de la primera solicitud, los parámetros ya están en la URL
    return cursos

# Función para obtener la página de inicio del curso
def obtener_pagina_inicio(curso_id):
    url = f"{CANVAS_API_URL}courses/{curso_id}/front_page"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json().get("body", "")
    return ""

# Función para revisar condiciones específicas en la página de bienvenida
def revisar_condiciones(pagina_inicio):
    check_div_style = '<div style="position: relative; width: 100%; color: white; overflow: hidden;">' in pagina_inicio

    # Analizar el HTML con BeautifulSoup
    soup = BeautifulSoup(pagina_inicio, 'html.parser')

    # Check de longitud de texto entre etiquetas específicas
    p_tag = soup.find('p', style=lambda value: value and 'text-align: justify;' in value)
    if p_tag:
        texto = p_tag.get_text(strip=True)
        check_texto_largo = len(texto) > 180
    else:
        check_texto_largo = False
    return check_div_style, check_texto_largo

# Función para revisar si hay un profesor registrado
def revisar_profesor(pagina_inicio):
    soup = BeautifulSoup(pagina_inicio, 'html.parser')
    p_tag = soup.find('p', style=lambda value: value and 'text-align: left; padding-left: 40px;' in value)
    if p_tag:
        strong_tag = p_tag.find('strong')
        if strong_tag:
            profesor = strong_tag.get_text(strip=True)
            return profesor != "Pendiente"
    return False

# Función para revisar el nombre técnico correcto
def revisar_nombre_tecnico(pagina_inicio):
    soup = BeautifulSoup(pagina_inicio, 'html.parser')
    strong_docente_en = soup.find('strong', string=lambda t: t and "Docente en:" in t)
    if strong_docente_en:
        p_docente_en = strong_docente_en.find_parent('p')
        if p_docente_en:
            ul = p_docente_en.find_next('ul')
            if ul:
                lis = ul.find_all('li')
                for li in lis:
                    texto_li = li.get_text(strip=True)
                    if texto_li:
                        return True
    return False

# Función para verificar las pestañas de navegación visibles
def verificar_pestanas_navegacion(curso_id):
    url = f"{CANVAS_API_URL}courses/{curso_id}/tabs"
    params = {
        "per_page": 100
    }
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        tabs = response.json()
        # Filtrar las pestañas visibles para los estudiantes
        pestañas_visibles = [tab['id'] for tab in tabs if tab.get('visibility') == 'public']
        # Pestañas esperadas (usando los IDs en inglés)
        pestañas_esperadas = ["home","modules", "grades", "people"]
        # Verificar si las pestañas visibles son exactamente las esperadas
        resultado = set(pestañas_visibles) == set(pestañas_esperadas)
        return resultado
    else:
        st.write(f"Error al obtener las pestañas del curso {curso_id}: {response.status_code}")
    return False

# Función para buscar "Programa de la asignatura" en los módulos
def buscar_programa_asignatura(curso_id):
    # Obtener los módulos del curso
    url = f"{CANVAS_API_URL}courses/{curso_id}/modules"
    params = {"per_page": 100}
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        modulos = response.json()
        for modulo in modulos:
            modulo_id = modulo['id']
            # Obtener los elementos del módulo
            url_items = f"{CANVAS_API_URL}courses/{curso_id}/modules/{modulo_id}/items"
            response_items = requests.get(url_items, headers=headers, params={"per_page": 100})
            if response_items.status_code == 200:
                items = response_items.json()
                for item in items:
                    if item['type'] == 'File' and item['title'] == "Programa de la asignatura":
                        # Obtener la información del archivo
                        file_id = item['content_id']
                        url_file = f"{CANVAS_API_URL}files/{file_id}"
                        response_file = requests.get(url_file, headers=headers)
                        if response_file.status_code == 200:
                            file_info = response_file.json()
                            file_url = file_info['url']
                            file_display_name = file_info.get('display_name', 'Archivo sin nombre')
                            if file_info.get('display_name') == 'Programa.pdf':
                                return '<span style="color:red;">❌</span>'
                            # Construir el enlace al archivo usando el nombre real del archivo
                            link_programa = f'<a href="{file_url}" target="_blank">{file_display_name}</a><span style="color:green;">✔️</span>'
                            return link_programa
    return '<span style="color:red;">❌</span>'

# Función para obtener las tareas del curso
def obtener_tareas_curso(curso_id):
    blocked_assigments = []
    url = f"{CANVAS_API_URL}courses/{curso_id}/assignments"
    params = {"per_page": 100}
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        tareas = response.json()
        if tareas:
            # Listar los nombres de las tareas
            nombres_tareas = [tarea['name'] for tarea in tareas if tarea['name'] not in blocked_assigments]
            if 'Tarea 1' in nombres_tareas:
                nombres_tareas.append('<span style="color:red;">❌</span>')
                return ", ".join(nombres_tareas)
            nombres_tareas.append('<span style="color:green;">✔️</span>')
            return ", ".join(nombres_tareas)
    return '<span style="color:red;">❌</span>'

# Función para convertir True/False a iconos de check y cruz
def convertir_icono(valor):
    if valor:
        return '<span style="color:green;">✔️</span>'
    else:
        return '<span style="color:red;">❌</span>'

# Interfaz de Streamlit
st.title("Revisión de cursos de Canvas LMS")

if subcuenta_seleccionada != "Seleccione una Especialidad":
    subcuenta_id = subcuentas[subcuenta_seleccionada]
    st.write(f"Obteniendo cursos de la subcuenta '{subcuenta_seleccionada}' (ID: {subcuenta_id})...")
    cursos = obtener_cursos(subcuenta_id)

    data = []
    for curso in cursos:
        nombre_curso = curso.get("name")
        curso_id = curso.get("id")
        sis_course_id = curso.get("sis_course_id", "N/A")
        # Construir el link al curso
        base_url = CANVAS_API_URL.replace('api/v1/', '')
        link_curso = f'<a href="{base_url}courses/{curso_id}" target="_blank">{nombre_curso}</a>'

        # Obtener la página de inicio del curso
        pagina_inicio = obtener_pagina_inicio(curso_id)

        # Revisar las condiciones
        check_div_style, check_texto_largo = revisar_condiciones(pagina_inicio)
        check_profesor = revisar_profesor(pagina_inicio)
        check_nombre_tecnico = revisar_nombre_tecnico(pagina_inicio)

        # Verificar las pestañas de navegación
        check_pestanas = verificar_pestanas_navegacion(curso_id)

        # Buscar "Programa de la asignatura"
        link_programa = buscar_programa_asignatura(curso_id)

        # Obtener las tareas del curso
        tareas_curso = obtener_tareas_curso(curso_id)

        data.append({
            "Nombre": link_curso,
            "id": curso_id,
            "sis_id": sis_course_id,
            "Bienvenida": convertir_icono(check_div_style),
            "Descripcion": convertir_icono(check_texto_largo),
            "Tutores": convertir_icono(check_profesor),
            "Nombre Tecnico": convertir_icono(check_nombre_tecnico),
            "Navegacion": convertir_icono(check_pestanas),
            "Programa": link_programa,
            "Tareas": tareas_curso
        })

    # Crear el DataFrame
    df = pd.DataFrame(data)

    # Ordenar las columnas si es necesario
    df = df[[
        "Nombre", "id", "sis_id",
        "Bienvenida", "Descripcion", "Tutores",
        "Nombre Tecnico", "Navegacion",
        "Programa", "Tareas"
    ]]

    # Mostrar la tabla con HTML
    st.write(df.to_html(escape=False, index=False), unsafe_allow_html=True)
    
    
    df_export = df.copy()

    # Eliminar etiquetas HTML de columnas específicas
    df_export['Nombre'] = df_export['Nombre'].str.replace(r'<.*?>', '', regex=True)
    df_export['Programa'] = df_export['Programa'].str.replace(r'<.*?>', '', regex=True)

    # Reemplazar iconos HTML por texto
    df_export = df_export.replace({
        '<span style="color:green;">✔️</span>': '✔️',
        '<span style="color:red;">❌</span>': '❌'
    }, regex=True)

    # Preparar el DataFrame para exportar
    column_settings = [
    {'width': 50},                           # Columna 1: Nombre del curso
    {'width': 12, 'alignment': 'center'},    # Columna 2: ID del curso
    {'width': 20},                           # Columna 3: SIS Course ID
    {'width': 15, 'alignment': 'center'},    # Columna 4: Div específico
    {'width': 13, 'alignment': 'center'},    # Columna 5: Texto > 200 caracteres
    {'width': 13, 'alignment': 'center'},    # Columna 6: Profesor registrado
    {'width': 15, 'alignment': 'center'},    # Columna 7: Nombre técnico correcto
    {'width': 13, 'alignment': 'center'},    # Columna 8: Pestañas de navegación correctas
    {'width': 56},                           # Columna 9: Programa de la asignatura
    {'width': 140}                           # Columna 10: Tareas del curso
    ]
    
    # Crear el archivo Excel en memoria
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df_export.to_excel(writer, index=False, sheet_name='Reporte')

    # Acceder al workbook y a la hoja
    workbook = writer.book
    worksheet = writer.sheets['Reporte']

    # Aplicar las configuraciones a las columnas
    for idx, column in enumerate(df_export.columns):
        column_letter = get_column_letter(idx + 1)
        settings = column_settings[idx]
        
        # Establecer el ancho de la columna
        worksheet.column_dimensions[column_letter].width = settings['width']
        
        # Aplicar alineación si está especificada
        if 'alignment' in settings:
            alignment = Alignment(horizontal=settings['alignment'])
            # Aplicar alineación a todas las celdas de la columna
            for cell in worksheet[column_letter]:
                cell.alignment = alignment
                
    # Aplicar colores al texto basado en el contenido para columnas 4 en adelante
    for col_idx in range(4, len(df_export.columns) + 1):
        column_letter = get_column_letter(col_idx)
        for cell in worksheet[column_letter][1:]:  # Excluir la primera fila (encabezados)
            if '✔️' in str(cell.value):
                cell.font = Font(color="008000")  # Color verde
            elif '❌' in str(cell.value):
                cell.font = Font(color="FF0000")  # Color rojo

    # Opcional: Centrar los encabezados
    for cell in worksheet[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Guardar el archivo Excel en memoria
    writer.close()
    processed_data = output.getvalue()

    st.download_button(label="Descargar Reporte",
                    data=processed_data,
                    file_name=f'{subcuenta_seleccionada}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # st.download_button(
    # label="Descargar Reporte",
    # data=processed_data,
    # file_name=f'{subcuenta_seleccionada}.xlsx',
    # mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    # )